import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def excelfilechange(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']


    a=sheet.max_row
    b=sheet.max_column

    for row in range(2, a+1):
        value=sheet.cell(row, 3)
        corrected_price=value.value*0.9
        new_cell = sheet.cell(row, 4)
        new_cell.value = corrected_price
        values= Reference(sheet, min_row=2,max_row= sheet.max_row,min_col=4,max_col=4)
        chart= BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'e2')
        new_filename=input('What is your new file"s name')
        wb.save(new_filename)



