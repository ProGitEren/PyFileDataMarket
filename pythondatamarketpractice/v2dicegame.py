import random
import openpyxl as xl

def control(a, b):
    print(f'Declined registerers: {a}')
    print(f'Entered registerers: {len(b)}')
    answer = input('Do you want to continue?(y for Yes,n for No): ')
    if answer == 'n':
        return False
    return True


def list_configure(filename):
    ##
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    maxirow = sheet.max_row
    maxicol = sheet.max_column
    minirow = sheet.min_row
    minicol = sheet.min_column
    ##

    playernumber = ''
    while type(playernumber) != int:
        ##
        try:
            playernumber = int(input('How many players are there: '))
        except Exception as err:
            print(err)
            print('Please enter a integer number!!!')
    list0 = []
    declined_registerer = 0
    while len(list0)<playernumber:
        registerer = input('What is the players name: ')
        registerer = registerer.lower()
        registerer = registerer.capitalize()
        if registerer in list0:
            print('This name is already in please enter a new name or enter with surname.')
            continue
        existence=False
        for row in range(minirow, maxirow):
            for col in range(minicol, maxicol):
                cell = sheet.cell(row, col)
                if cell.value == registerer:
                    existence=True

        if not existence:
            print('This player name is not on the list. Please choose another person')
            declined_registerer += 1
            Answerr = control(declined_registerer, list0)
            if not Answerr:
                break
            continue
        list0.append(registerer)
        Answer = control(declined_registerer, list0)
        if not Answer:
            break


    return list0


def dice_game(players):
    list_of_numbers = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
    ##
    dictofplayers = {}
    registeredpeople = []
    ##

    for i in range(len(players)):
        player = ''
        while player not in players or player in registeredpeople:
            player = input('Which one is going to roll the dices? ')
            player = player.lower()
            player = player.capitalize()
            if player in registeredpeople:
                print('This player is already registered. Please enter a new name or enter with surname.')

        ##
        registeredpeople.append(player)
        first = random.choice(list_of_numbers)
        second = random.choice(list_of_numbers)
        sum = first + second
        print(f'Player {player} has the sum {sum}')
        dictofplayers[player] = sum

    ##

    listofsums = list(dictofplayers.values())
    maximum = max(listofsums)
    winner = 0
    for i in dictofplayers.keys():
        if dictofplayers[i] == maximum:
            winner += 1
            print(f'Player {i} has won the game with the sum {maximum}')
    if winner > 1:
        print('More than one player has won the game no absolute winners')


Players_list = list_configure('listpeople.xlsx')
dice_game(Players_list)